import imaplib, email, os, re, csv, smtplib, pytz, zipfile, requests, error #type: ignore
import pandas as pd #type: ignore
from datetime import datetime
from openpyxl import load_workbook #type: ignore
from email.message import EmailMessage
from email.header import decode_header
from bs4 import BeautifulSoup #type: ignore
from dotenv import load_dotenv #type: ignore
from email import utils

load_dotenv(override=True)

email_address = os.getenv("EMAIL_ADDRESS")
email_password = os.getenv("EMAIL_PASSWORD")
email_from_1 = os.getenv("EMAIL_FROM_1")
email_from_2 = os.getenv("EMAIL_FROM_2")
recipient_1 = os.getenv("RECIPIENT_1")
recipient_2 = os.getenv("RECIPIENT_2")
excel_file_path = os.getenv("EXCEL_FILE_PATH")
tsv_file_path = os.getenv("TSV_FILE_PATH")
sheet_name = os.getenv("SHEET_NAME")
error_excel_path = os.getenv("ERROR_EXCEL_PATH")
shipping_txt_file = os.getenv("SHIPPING_TXT_FILE")
walmart_order_excel_file = os.getenv("WALMART_ORDER_EXCEL_FILE")

recipients = [recipient_1, recipient_2]

ebay_tracking_pattern = re.compile(r'Tracking number\s*:\s*(\S+)', re.IGNORECASE)
ebay_order_pattern = re.compile(r'^\s*\d{2}-\d{5}-\d{5}\s*$')

keurig_tracking_pattern = re.compile(r'Tracking\s*#\s*:\s*(\S+)', re.IGNORECASE)
keurig_order_pattern = re.compile(r'Order\s*#\s*:\s*(\S+)', re.IGNORECASE)

mail = imaplib.IMAP4_SSL("imap.gmail.com")
smtp = smtplib.SMTP_SSL('smtp.gmail.com', 465)

mail.login(email_address, email_password)
smtp.login(email_address, email_password)
local_tz = pytz.timezone('America/New_York')

headers = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.10 Safari/605.1.1",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "en-US,en;q=0.5",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Cache-Control": "max-age=0"
}

soup = None

def get_carrier(tracking_number):
    if tracking_number.startswith('1Z'):
        return 'UPS'
    elif len(tracking_number) in (15, 12):
        return 'FedEx'
    elif tracking_number.startswith('92'):
        return 'USPS'
    
def scrape_tracking_link(soup):
    for a in soup.find_all('a'):
        a_text = a.get_text()
        a_pattern = re.compile(r'\btrack (order|delivery)\b', re.IGNORECASE)
        found_a = re.findall(pattern=a_pattern, string=a_text) 
        if found_a:
            href  = a.get('href')
            return href
        
def scrape_pattern_in_email(html_soup, element, pattern):
    for e in html_soup.find_all(element):
        e_text = e.get_text()
        found_e = re.findall(pattern=pattern, string=e_text)
        if len(found_e) > 0:
            return found_e
        
def mark_email_as_trash(email_id, mail=mail):
    mail.store(email_id, '+X-GM-LABELS', '\\Trash')

def send_message(subject: str, email_msg, email_=email_address, recipients=recipients):
    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = email_
    msg['To'] = ', '.join(recipients)
    msg.set_content(email_msg)
    msg.add_alternative(email_msg, subtype='html')
    smtp.send_message(msg=msg)

    
try:
    print('Login Successful')
    mail.select('INBOX')
    status, result = mail.search(None, f'OR FROM {email_from_1} FROM {email_from_2}')

    email_ids = result[0].split()
    for i in range(len(email_ids)):
        try:
            full_address = None
            order = []
            tracking = []
            status, msg_data = mail.fetch(email_ids[i], '(RFC822)')
            
            raw_email = email.message_from_bytes(msg_data[0][1])

            subject_, encoding = decode_header(raw_email['Subject'])[0]
            date_string = raw_email.get('Date')

            email_date = utils.parsedate_to_datetime(date_string).astimezone(local_tz).replace(tzinfo=None)
                
            
            for part in raw_email.walk():
                content_type = part.get_content_type()
            
                if content_type == 'text/html':
                    pl = part.get_payload(decode=True)
                    soup = BeautifulSoup(pl, 'html.parser')
                    
                    # Get tracking link from email
                    href = scrape_tracking_link(soup=soup)
                    
                    # Get tracking number for eBay 
                    tracking = scrape_pattern_in_email(soup, 'p', ebay_tracking_pattern)
                    # Get order number for eBay 
                    order = scrape_pattern_in_email(soup, 'span', ebay_order_pattern)

                    # If no eBay tracking number 
                    if tracking is None:
                        # Get Keurig tracking number 
                        tracking = scrape_pattern_in_email(soup, 'td', keurig_tracking_pattern)
                        if tracking is None:
                            pass
                    
                    # If no eBay order number 
                    if order is None:
                        # Get Keurig order number 
                        order = scrape_pattern_in_email(soup, 'td', keurig_order_pattern)
                
                    # ebay and keurig shipping address scraping 
                    shipping_td = soup.find('td', string=lambda t: t and 'Shipping Address' in t)
                    e_shipping_h3 = soup.find('h3', string=lambda text: text and 'Your order will ' in text)  
                    # Keurig shipping address scrape
                    if shipping_td:
                        parent_tr = shipping_td.find_parent('tr')
                        address_table = parent_tr.find_parent('table')
                        all_rows = address_table.find_all('tr')
                        start_index = all_rows.index(parent_tr) + 1
                        address_lines = []

                        for row in all_rows[start_index:]:
                            tds = row.find_all('td')
                            for td in tds:
                                text = td.get_text()
                                text = text.lstrip()
                                if text not in address_lines:
                                    address_lines.append(text)

                        address_lines = [item for item in address_lines if item != '']
                        full_address = "\t".join(address_lines)
                    elif e_shipping_h3:
                        shipping_p = e_shipping_h3.find_next_sibling('p')
                        if shipping_p:
                            full_address = shipping_p.get_text(separator='\t').strip()

            # If no tracking in email and there is a link to ebay tracking 
            #  scrape tracking number from ebay website with provided link to track order
            if tracking is None and soup is not None:
                try:
                    tracking = []
                    url = scrape_tracking_link(soup=soup)

                    rspn = requests.get(url=url, headers=headers, allow_redirects=True)

                    if rspn.status_code != 200:
                        raise error.No_Tracking_Number(f'<html><p>No tracking number found in email {i}<br /> where customer shipping address is: {full_address}<br /> and the order number is {order}<br /><br /><br />P.S. There might be more issues with this email</p><a href="{href}">Track Order</a>')
                        
                    ebay_soup = BeautifulSoup(rspn.text, 'html.parser')

                    found_ = ebay_soup.find('span', string=lambda t: t and 'Number' in t)

                    parent_dt = found_.find_parent()
                    parent_div = parent_dt.find_parent()

                    spans = parent_div.find_all('span')
                    tracking.append(spans[-1].get_text())

                    mark_email_as_trash(mail=mail, email_id=email_ids[i])

                    if tracking is None or len(tracking[0]) < 10 :
                        raise error.No_Tracking_Number(f'<html><p>No tracking number found in email {i}<br /> where customer shipping address is: {full_address}<br /> and the order number is {order}<br /><br /><br />P.S. There might be more issues with this email</p><a href="{href}">Track Order</a>')
                        
                except requests.exceptions.RequestException as e:
                    print(f"Request failed: {e}")        
            elif order is None:
                raise error.No_Order_Number(f'<html><p>No order number found in email {i}<br />where customer shipping address is: {full_address}<br />and tracking number is {tracking}<br /><br /><br />P.S. There might be more issues with this email</p><a href="{href}">Track Order</a></html>')
            elif full_address == None or full_address == '':
                raise error.No_Shipping_Address(f'<html><p>No shipping address found in email {i}<br />where order number is: {order}<br />and tracking number is {tracking} as a result can not find the amazon order number<br /><br /><br /> P.S. There might be more issues with this email</p><a href="{href}">Track Order</a>')
            
            zip_code_pattern = re.compile(r'\b(\d{5})(?:-\d{4})?\b')
            zip_code = re.findall(pattern=zip_code_pattern, string=full_address)
            address = re.split(r'\t+', full_address)
            name = re.sub(r'\s+', ' ', address[0]).strip()
                

            zip = zip_code[-1]
            found_match_amazon = False
            found_match = False
            try:
                with open(tsv_file_path, 'r') as file:
                    reader = csv.reader(file, delimiter='\t')
                    for row_n, row in enumerate(list(reader)):
                        if len(row) > 0:
                            if name.lower() in row[17].lower().replace('.', '') and zip in row[23]:
                                    
                                found_match_amazon = True
                                a_order_id = row[0]

                                data = []
                                carrier = get_carrier(tracking_number=tracking[0])
                                ship_date = datetime.now().date()
                                
                                data.append([ship_date, tracking[0], a_order_id, carrier])

                                try:
                                    wb = load_workbook(excel_file_path)
                                    sheet = wb[sheet_name]

                                    max_row = sheet.max_row

                                    for row_num, data_to_append in enumerate(data, start=max_row + 1):
                                        sheet.cell(row=row_num, column=1, value=data_to_append[2])
                                        sheet.cell(row=row_num, column=7, value=data_to_append[1])
                                        sheet.cell(row=row_num, column=5, value=data_to_append[3])
                                        sheet.cell(row=row_num, column=4, value=data_to_append[0])

                                    wb.save(excel_file_path)
                                        
                                    mark_email_as_trash(email_id=email_ids[i])
                                    del data
                                    break
                                except FileNotFoundError:
                                    print(f'Error: No file found at: {excel_file_path}')
                                except PermissionError:
                                    print(f'Error: Permission denied most probably cause file open in another program, close file, and try again')
                                except zipfile.BadZipFile:
                                    print(f'BadZipFile caught file at {excel_file_path} is not a valid .xlsx (Excel) file')
            except FileNotFoundError:
                print(f'Error: No file found at: {tsv_file_path}')
            except PermissionError:
                print(f'Error: Permission denied to open file at: {tsv_file_path}')
            except OSError as e:
                print(f'An unexpected error occured: {e}')
            if not found_match_amazon:
                try:
                    wm_wb = load_workbook(walmart_order_excel_file)
                    wm_ws = wm_wb['Po Details']

                    carrier = get_carrier(tracking_number=tracking[0])

                    for row_id, row in enumerate(wm_ws.iter_rows(values_only=True), start=1):
                        cleaned_name = re.sub(r'\s+', ' ', row[5])
                        if zip == row[13] and name == cleaned_name:
                            found_match = True

                            wm_ws.cell(row=row_id, column=32, value=carrier)
                            wm_ws.cell(row=row_id, column=33, value=tracking[0]) 
                            mark_email_as_trash(email_id=email_ids[i])

                    wm_wb.save(walmart_order_excel_file)     
                except FileNotFoundError:
                    print(f'Error: No file found at: {excel_file_path}')
                except PermissionError:
                    print(f'Error: Permission denied most probably cause file open in another program, close file, and try again')
                except zipfile.BadZipFile:
                    print(f'BadZipFile caught file at {excel_file_path} is not a valid .xlsx (Excel) file')
                except Exception as e:
                    print(f'Unexpected error: {e}')
            if not found_match:
                error_message = 'Did not find match in walmart nor amazon'
                try:
                    error_wb = load_workbook(error_excel_path)
                    error_sheet = error_wb['Sheet1']
                    error_max_row = error_sheet.max_row

                    error_data = []

                    error_data.append([error_message, order[0], tracking[0], full_address, name, zip, email_date, datetime.now()])

                    for error_row_num, error_row_data in enumerate(error_data, start=error_max_row+1):
                        error_sheet.cell(row=error_row_num, column=1, value=error_row_data[0])
                        error_sheet.cell(row=error_row_num, column=2, value=error_row_data[1])
                        error_sheet.cell(row=error_row_num, column=3, value=error_row_data[2])
                        error_sheet.cell(row=error_row_num, column=4, value=error_row_data[3])
                        error_sheet.cell(row=error_row_num, column=5, value=error_row_data[4])
                        error_sheet.cell(row=error_row_num, column=6, value=error_row_data[5])
                        error_sheet.cell(row=error_row_num, column=7, value=error_row_data[6])
                        error_sheet.cell(row=error_row_num, column=8, value=error_row_data[7])

                    error_wb.save(error_excel_path)
                except FileNotFoundError:
                    print(f'Error: No file found at: {error_excel_path}')
                except PermissionError:
                    print(f'Error: Permission denied most probably cause file open in another program, close file, and try again')
                except zipfile.BadZipFile:
                    print(f'BadZipFile caught file at {error_excel_path} is not a valid .xlsx (Excel) file')
                except Exception as e:
                    print(f'Unexpected error: {e}')
        except error.No_Order_Number as non_e:
            send_message('No Order Number', str(non_e))
            mark_email_as_trash(email_id=email_ids[i])
        except error.No_Tracking_Number as ntn_e:
            send_message('No Tracking Number', str(ntn_e))
            mark_email_as_trash(email_id=email_ids[i])
        except error.No_Shipping_Address as nsa_e:
            send_message('No Shipping Address', str(nsa_e))
            mark_email_as_trash(email_id=email_ids[i])
        except Exception as e:
            print(f'Unexpected error: {e}')
        print(f'----------Processed email #{i}---------')
    mail.close()
    mail.logout()

    print("logged out succefully")
    try:
        ef = pd.read_excel(excel_file_path, engine='openpyxl', sheet_name=sheet_name)
        ef.to_csv(shipping_txt_file, sep='\t', index=False)
    except FileNotFoundError:
        print(f'Error: No file found at {shipping_txt_file}')
    except Exception as e:
        print(f'There was an error converitng excel file to txt tab delimited file: {e}')
        
except imaplib.IMAP4_SSL.error as mail_s_e:
    print(f'error: {mail_s_e}')
except Exception as e:
    print(f'Unexpected error: {e}')