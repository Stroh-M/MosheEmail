from openpyxl import load_workbook # type:ignore



wb = load_workbook("C:\\Users\\meir.stroh\\OneDrive\\MosheEmail\\WalmartOrders.xlsx")
ws = wb['Po Details']

max_rows = ws.max_row

for row_id, row in enumerate(ws.iter_rows(values_only=True)):
    if '49091' == row[13]:
        ws.cell(row=row_id+1, column=32, value='done')
        print('matches')
        print(row[13])


    # print(row_id)
    # print(row[0])
    
wb.save("C:\\Users\\meir.stroh\\OneDrive\\MosheEmail\\WalmartOrders.xlsx")