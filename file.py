from openpyxl import load_workbook #type: ignore 

file_path = r'C:\\Users\\meir.stroh\\OneDrive\\MosheEmail\\Flat.File.ShippingConfirm (1).xlsx'

# Load workbook
wb = load_workbook(file_path)

# List available sheet names (for debugging)
print("Available sheets:", wb.sheetnames)

order_number = 67847346292349
tracking_number = 634926538593480
carrier = 'UPS'

data = []
a = True
b = 0
while a:
    data.append([order_number, tracking_number, carrier])
    order_number = order_number + 10
    tracking_number = tracking_number + 10
    b += 1 
    if b == 10:
        a = False


print(data)


ws = wb['ShippingConfirmation']

# for i, row_data in enumerate(data, start=2):
ws.cell(row=2, column=1, value='hey')
    # print(i)
    # print(row_data)






# import csv
# # from openpyxl import load_workbook

# with open('C:\\Users\\meir.stroh\\OneDrive\\MosheEmail\\118545073774020185.txt', 'r') as file:
#     reader = csv.reader(file, delimiter='\t')
#     rows = list(reader)

#     # print(rows)
#     for row in rows:
#         if '114-5883196-2075457' in row:
#             print("found")

#     for x, elem in enumerate(rows):
#         if '114-5883196-2075457' in elem:
#             print(x)

#             rows[x][1] = 'Rosenberg'

#         with open('C:\\Users\\meir.stroh\\OneDrive\\MosheEmail\\118545073774020185.txt', 'w', newline='') as file:
#             writer = csv.writer(file, delimiter='\t')
#             writer.writerows(rows)


    

        